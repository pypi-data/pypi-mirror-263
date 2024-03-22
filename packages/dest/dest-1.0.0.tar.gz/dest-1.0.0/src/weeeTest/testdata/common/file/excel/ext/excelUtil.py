import asyncio
import os.path
import time

import aiofiles
from aiocsv import AsyncReader
from openpyxl import load_workbook


class ExcelUtil:

    @staticmethod
    async def _excel_read(file_path: str = None, file_name: str = None, return_type: str = None, sheet: str = "Sheet1",
                          line: int = 1, end_line: int = None) -> list:
        """
        Convert Excel file data to list
        :param file_path: Path to file
        :param sheet: Excel sheet, default name is Sheet1
        :param line: Start line of read data
        :param end_line: Start line of read data
        :return: list data

        Usage:
            excel_to_list("data.xlsx", sheet="Sheet1", line=1)
        """
        path = os.path.join(file_path, file_name)
        async with aiofiles.open(path, 'rb+') as file:
            await file.read()
            excel_table = load_workbook(path)
            try:
                if file_name is None:
                    raise Exception('文件名不能为空')

                is_contain = file_name.endswith('.xlsx')
                if not is_contain:
                    file_name += '.xlsx'
                sheet = excel_table[sheet]
                if end_line is None:
                    end_line = sheet.max_row

                if return_type is None or 'list' in return_type:
                    table_data = []
                    for i in sheet.iter_rows(line, end_line):
                        line_data = []
                        for field in i:
                            line_data.append(field.value)
                        table_data.append(line_data)
                    return table_data
                else:
                    raise Exception('excel读取只支持返回list')
            except FileNotFoundError:
                raise Exception(f"文件不存在，请确认file_path:{os.path.abspath(path)}")

    @staticmethod
    async def _csv_read(file_path: str = None, file_name: str = None, return_type: str = None, line: int = 1,
                        end_line: int = None) -> list:
        """
        Convert CSV file data to list
        :param file_path: Path to file
        :param line: Start line of read data
        :param end_line: End line of read data
        :return: list data

        Usage:
            csv_to_list("data.csv", line=1)
        """
        try:
            if file_name is None:
                raise Exception('文件名不能为空')
            else:
                is_contain = file_name.endswith('.csv')
                if not is_contain:
                    file_name += '.csv'
            if return_type is None or 'list' in return_type:
                table_data = []
                async with aiofiles.open(os.path.join(file_path, file_name), 'r', encoding='utf_8_sig') as csv_file:
                    async for row in AsyncReader(csv_file):
                        table_data.append(row)
                return table_data
            else:
                raise Exception('csv读取只支持返回list')
        except FileNotFoundError:
            raise Exception(f"文件不存在，请确认file_path:{os.path.abspath(os.path.join(file_path, file_name))}")
        except RuntimeError as e:
            raise Exception(str(e))


if __name__ == '__main__':
    data_list = asyncio.run(
        ExcelUtil._excel_read(file_path='D:\\py_project_company\\qa-weeetest-sdk\\weeeTest\\demo\\test_data',
                              file_name='excel_data.xlsx'))
    print(data_list)

    # data_list = asyncio.run(
    #     ExcelUtil._csv_read(file_path='D:\\py_project_company\\qa-weeetest-sdk\\weeeTest\\demo\\test_data',
    #                         file_name='csv_data', return_type='list'))
    # print(data_list)

    # data_list = ExcelUtil._excel_read(file_path='./data/data/testData', file_name='excel_data.xlsx')
    # print(data_list)
