from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import pandas as pd
from datetime import datetime
from pandas.api.types import is_float_dtype, is_integer_dtype
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule

class ExcelWriter():
    def __init__(self, sheet_name = "Ana Sayfa"):
        self.title_border = Border(
            top=Side(border_style="thin",color='FF000000'),
            bottom=Side(border_style="thin",color='FF000000'))

        self.border = Border(
            top=Side(border_style="thin",color='FF000000'),
            bottom=Side(border_style="thin",color='FF000000'))

        self.title_font = Font(
            name='Calibri',
            color='FFFFFFFF',
            size=10,
            bold=True,
            italic=False)
        self.font = Font(
            name='Calibri',
            size=10,
            bold=False,
            italic=False)

        self.number_format = "#,##0.00"
        self.integer_format = "0"
        self.wb = Workbook()

        self.ws0 = self.wb[self.wb.worksheets[0].title]
        self.ws0.title = sheet_name
        self.page_no = 0
        self.table_no = 0
        self.ws0.cell(1,1).value = "Oluşturulma Tarihi"
        self.ws0.cell(2,1).value = datetime.now()


    def create_worksheet(self, sheet_name = "Veri"):
        self.page_no += 1
        if sheet_name == "Veri":
            name = sheet_name + str(self.page_no)
        else:
            name = sheet_name
            
        self.wb.create_sheet(name)
        setattr(self, "ws"+str(self.page_no), self.wb[name])

    def write_dataframe(self, ws_no, df, 
                        start_cell_row = 1, start_cell_col = 1):
        ws = getattr(self, "ws"+str(ws_no))
        self.table_no += 1

        ref = self.get_reference(start_cell_row, start_cell_col, df.shape)
        i,j = df.shape
        for n in range(j):
            ws.cell(start_cell_row, start_cell_col + n).value = str(df.columns[n])
            self.apply_style(ws_no, start_cell_row, start_cell_col + n, df.columns[n], title =True)
            for m in range(i):
                ws.cell(start_cell_row + m + 1, start_cell_col + n).value = df.iloc[m,n]
                self.apply_style(ws_no, start_cell_row + m + 1, start_cell_col + n, df.iloc[m,n], 
                                 title =False)
                
                
        tab = Table(displayName="Table"+str(self.table_no), ref=ref)
        style = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)#showFirstColumn=False,showLastColumn=False, showRowStripes=True, showColumnStripes=True
        tab.tableStyleInfo = style
        ws.add_table(tab)

    def apply_style(self, ws_no, row, col, val, title =False):
        ws = getattr(self, "ws"+str(ws_no))
        if title:
            ws.cell(row, col).font = self.title_font
            ws.cell(row, col).border = self.title_border
            ws.cell(row, col).alignment = Alignment(horizontal='center', vertical='center')
        else:
            ws.cell(row, col).font = self.font
            ws.cell(row, col).border = self.border
            ws.cell(row, col).alignment = Alignment(horizontal='center' if type(val) == str else "right",
                                                         vertical='center')            
            if is_integer_dtype(val):
                ws.cell(row, col).number_format = self.integer_format
            else:
                if is_float_dtype(val):
                    ws.cell(row, col).number_format = self.number_format
                else:
                    pass
            
    def adjust_columns(self, width_exp = 1.4):
        for k in [i for i in dir(self) if i[:2] == "ws"]:
            ws = getattr(self, k)
            dim_holder = DimensionHolder(worksheet=ws)

            for col in range(ws.min_column, ws.max_column + 1):
                max_dimension = 2
                for row in range(1, ws.max_row + 1):
                    try:
                        val = len(str(round(ws.cell(row,col).value,2)))
                    except:
                        val = len(str(ws.cell(row,col).value))

                    if val > max_dimension:
                        max_dimension = val
                ws.column_dimensions[get_column_letter(col)].width = max_dimension*width_exp

    def apply_color_scale(self, ws_no):
        ws = getattr(self, "ws"+str(ws_no))
#        for nth,i in enumerate([aaa.change_month, aaa.change_week, aaa.change_day]):
            # vv = getattr(ex, "ws{:.0f}".format(nth+2))
        # for c in [get_column_letter(n+1) for n,j in enumerate(i.columns) if j[-4:] == "P(%)"]:
        for c in [get_column_letter(n) for n in range(ws.min_column, ws.max_column + 1)]:
            ws.conditional_formatting.add(
                '{}{:.0f}:{}{:.0f}'.format(c, ws.min_row, c, ws.max_row + 1),
                ColorScaleRule(
                    start_type='min', start_color='F8696B',
                    mid_type='percent', mid_color='FFEB84',
                    end_type='max', end_color='63BE7B'
                    )
                )


    @staticmethod
    def get_reference(start_cell_row, start_cell_col, shape):
        row,col = shape
        ref = "{}:{}".format(get_column_letter(start_cell_col) + str(start_cell_row),
                             get_column_letter(start_cell_col+col-1) + str(start_cell_row+row))
        return ref


